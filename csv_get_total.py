##########################################################################

import glob
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import messagebox as tmsg

root = tk.Tk()
root.withdraw()

# csvが格納されているフォルダを選択
target_path = r"C:\Users\name\OneDrive\デスクトップ\pandas\*.csv"

file_paths = glob.glob(target_path)
# print(file_paths)

# 最終結果を更新するbookのパス
write_wb = openpyxl.load_workbook(r"C:\Users\name\OneDrive\デスクトップ\pandas\ans.xlsx")
write_ws = write_wb["Sheet1"]

files_count = len(file_paths)
total_all = []
total_d = []
total_p = []

for i in file_paths:
    df = pd.read_csv(i, encoding='shift_jis')
    df_all = pd.DataFrame(df["USD JPY"])
    df_ans_all = df_all.sum()
    total_all.append(df_ans_all.iloc[-1])

    df_d = pd.DataFrame(df[(df["状態"] == "YES") & (df["No"] == 111)])
    df_d = pd.DataFrame(df_d["USD JPY"])
    df_ans_d = df_d.sum()
    total_d.append(df_ans_d.iloc[-1])

    df_p = pd.DataFrame(df[(df["状態"] == "YES") & (df["No"] == 222)])
    df_p = pd.DataFrame(df_p["USD JPY"])
    df_ans_p = df_p.sum()
    total_p.append(df_ans_p.iloc[-1])
    print(i)

all_max = write_ws.cell(1, 2)
all_average = write_ws.cell(1, 3)

all_max.value = max(total_all)
all_average.value = sum(total_all) / files_count

d_max = write_ws.cell(2, 2)
d_average = write_ws.cell(2, 3)

d_max.value = max(total_d)
d_average.value = sum(total_d) / files_count

p_max = write_ws.cell(3, 2)
p_average = write_ws.cell(3, 3)

p_max.value = max(total_p)
p_average.value = sum(total_p) / files_count

csv_count_output = write_ws.cell(4, 1)
csv_count_output.value = files_count

write_wb.save(r"C:\Users\name\OneDrive\デスクトップ\pandas\ans.xlsx")
tmsg.showinfo('完了報告', '処理が終了しました！！')

##########################################################################
