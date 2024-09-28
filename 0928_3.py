import openpyxl
# fname = r"C:\Users\turnt\OneDrive\デスクトップ\Python\2024_python_practice\sample00001.xlsx"
#
# wb = openpyxl.load_workbook(fname)
# sheet = wb.active
#
# print(sheet.title)

wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet()
wb.create_sheet()
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=2, title="test001")
del wb["Sheet2"]
wb.active = wb["test001"]
wb.save("excel002.xlsx")