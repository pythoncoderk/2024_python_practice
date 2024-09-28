import openpyxl

fname = r"C:\Users\turnt\OneDrive\デスクトップ\Python\2024_python_practice\sample00001.xlsx"

wb = openpyxl.load_workbook(fname)
sheet = wb.active

print(sheet.title)
print(sheet.max_row)
print(sheet.max_column)

sheet2 = wb.copy_worksheet(sheet)
sheet2 = wb.copy_worksheet(sheet)
sheet2 = wb.copy_worksheet(sheet)
sheet2 = wb.copy_worksheet(sheet)

sheet2.title = sheet.title + " - コピー"
sheet2.sheet_properties.tabColor = "1072ba"
print(wb.sheetnames)
wb.save("tmp.xlsx")


wb = openpyxl.load_workbook(r"C:\Users\turnt\OneDrive\デスクトップ\Python\2024_python_practice\excel003.xlsx")
sheet = wb.active
print(sheet.max_row)
print(sheet.max_column)
sheet.insert_rows(5, 5)
sheet.insert_cols(5, 5)
wb.save("tmp2.xlsx")