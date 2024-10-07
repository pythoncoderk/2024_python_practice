import openpyxl

fname = r"C:\Users\turnt\OneDrive\デスクトップ\Python\2024_python_practice\sample00001.xlsx"

wb = openpyxl.load_workbook(fname, data_only=True)
wb.save("out.xlsx")

for sheet in wb:
    print(sheet.title)

sheetnames = wb.sheetnames
print(sheetnames)